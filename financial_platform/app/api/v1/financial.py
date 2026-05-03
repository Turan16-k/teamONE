"""
T3 + T10: Finansal rapor CRUD endpoint'leri.
Güvenlik: magic byte doğrulaması, dosya boyutu, Content-Disposition sanitizasyonu,
para birimi whitelist, sayfalama sınırları.
"""
import csv
import io
import os
import re
import unicodedata
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, status, Request, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.api.deps import get_current_active_user
from app.models.user import User
from app.models.company import Company
from app.models.financial import FinancialReport, ReportType, PeriodType
from app.models.log import LogAction
from app.schemas.financial import FinancialReportCreate, FinancialReportUpdate, FinancialReportResponse
from app.core.rbac import require_owner_or_admin
from app.services.ai_service import ai_service
from app.services.pptx_service import pptx_service
from app.services import report_generator
from app.models.extended import CompanyBank, Collection, CompanyProject, Investment
from app.services.notification_service import notify_user_report_ready
from app.utils.logging import log_audit, log_exception
from app.utils.pagination import PaginationParams, paginate
from app.models.subscription import UserSubscription, SubscriptionStatus

router = APIRouter(prefix="/financial", tags=["Financial Reports"])

# ── Sabitler ──────────────────────────────────────────────────────────────────
_MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB
_CURRENT_YEAR = datetime.utcnow().year
_ALLOWED_CURRENCIES = {"TRY", "USD", "EUR", "GBP"}

ALLOWED_MEDIA_TYPES = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "image/webp",
}


# ── Yardımcı Fonksiyonlar ─────────────────────────────────────────────────────

def _check_and_increment_ai_quota(user: User, db: Session) -> None:
    sub: Optional[UserSubscription] = user.subscription
    if not sub or sub.status != SubscriptionStatus.ACTIVE:
        raise HTTPException(
            status_code=402,
            detail="Aktif bir aboneliğiniz yok. Lütfen bir paket satın alın.",
        )
    pkg = sub.package
    if pkg and sub.ai_calls_used >= pkg.max_ai_calls_per_month:
        raise HTTPException(
            status_code=429,
            detail=f"Aylık AI çağrı limitine ulaştınız ({pkg.max_ai_calls_per_month} çağrı).",
        )
    sub.ai_calls_used += 1
    db.flush()


def _get_company_or_403(company_id: int, current_user: User, db: Session) -> Company:
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Şirket bulunamadı.")
    require_owner_or_admin(company.owner_id, current_user)
    return company


def _validate_magic_bytes(content: bytes) -> str:
    """İçeriğin magic bytes'ına göre gerçek MIME türünü döner; geçersizse 400 fırlatır."""
    if content[:4] == b"%PDF":
        return "application/pdf"
    if content[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if content[:8] == b"\x89PNG\r\n\x1a\n":
        return "image/png"
    if content[:4] == b"RIFF" and content[8:12] == b"WEBP":
        return "image/webp"
    raise HTTPException(
        status_code=400,
        detail="Geçersiz dosya içeriği. PDF, PNG, JPG veya WEBP yükleyin.",
    )


def _sanitize_filename(name: str) -> str:
    """Header injection ve path traversal karakterlerini temizler."""
    # Unicode normalize, sadece ASCII harfler/rakamlar/tire/alt çizgi/nokta bırak
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    name = re.sub(r"[^\w\-.]", "_", name).strip("._")
    return name[:100] or "rapor"


def _safe_cd_header(filename: str) -> str:
    """RFC 5987 uyumlu Content-Disposition header değeri üretir."""
    safe = _sanitize_filename(filename)
    return f'attachment; filename="{safe}"'


def _validate_fiscal_year(year: int) -> None:
    if not (1900 <= year <= _CURRENT_YEAR + 1):
        raise HTTPException(status_code=400, detail=f"Geçersiz mali yıl: {year}.")


def _safe_store_filename(raw: Optional[str]) -> Optional[str]:
    """Kullanıcıdan gelen dosya adını DB'ye yazmadan önce sanitize eder."""
    if not raw:
        return None
    base = os.path.basename(raw.replace("\\", "/"))
    base = base.replace("\x00", "").strip()
    return base[:255] or None


# ── Endpoint'ler ──────────────────────────────────────────────────────────────

@router.get("/companies/{company_id}/reports", response_model=dict)
def list_reports(
    company_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    fiscal_year: Optional[int] = Query(None, ge=1900, le=2100),
    report_type: Optional[str] = Query(None),
    period: Optional[str] = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> dict:
    _get_company_or_403(company_id, current_user, db)
    query = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.company_id == company_id)
    )
    if fiscal_year:
        query = query.filter(FinancialReport.fiscal_year == fiscal_year)
    if report_type:
        try:
            query = query.filter(FinancialReport.report_type == ReportType(report_type))
        except ValueError:
            pass
    if period:
        try:
            query = query.filter(FinancialReport.period == PeriodType(period))
        except ValueError:
            pass
    query = query.order_by(FinancialReport.fiscal_year.desc(), FinancialReport.created_at.desc())
    result = paginate(query, PaginationParams(page=page, page_size=page_size))
    result["items"] = [FinancialReportResponse.model_validate(r) for r in result["items"]]
    return result


@router.post("/companies/{company_id}/reports",
             response_model=FinancialReportResponse, status_code=status.HTTP_201_CREATED)
def create_report(
    company_id: int,
    payload: FinancialReportCreate,
    request: Request,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> FinancialReport:
    _get_company_or_403(company_id, current_user, db)
    _validate_fiscal_year(payload.fiscal_year)
    payload.company_id = company_id

    report = FinancialReport(**payload.model_dump())
    db.add(report)
    db.flush()
    log_audit(db, current_user.id, LogAction.CREATE, "FinancialReport", report.id,
              new_values={"company_id": company_id, "fiscal_year": payload.fiscal_year,
                          "report_type": payload.report_type.value},
              ip_address=request.client.host if request.client else None)
    db.commit()
    db.refresh(report)
    return report


@router.get("/reports/{report_id}", response_model=FinancialReportResponse)
def get_report(
    report_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> FinancialReport:
    report = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı.")
    require_owner_or_admin(report.company.owner_id, current_user)
    return report


@router.put("/reports/{report_id}", response_model=FinancialReportResponse)
def update_report(
    report_id: int,
    payload: FinancialReportUpdate,
    request: Request,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> FinancialReport:
    """T3: Çift yönlü veri bağlama — AI doldurdu, kullanıcı manuel düzeltiyor."""
    report = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı.")
    require_owner_or_admin(report.company.owner_id, current_user)

    changed = payload.model_dump(exclude_none=True)
    old_vals = {k: str(getattr(report, k)) for k in changed if getattr(report, k) is not None}
    for field, value in changed.items():
        setattr(report, field, value)

    log_audit(db, current_user.id, LogAction.UPDATE, "FinancialReport", report_id,
              old_values=old_vals, new_values={k: str(v) for k, v in changed.items()},
              ip_address=request.client.host if request.client else None)
    db.commit()
    db.refresh(report)
    return report


@router.post("/reports/{report_id}/analyze", response_model=dict)
def trigger_ai_analysis(
    report_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> dict:
    """T3: Analitik Modül — LLM ile finansal rasyo ve değerlendirme üretir."""
    report = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı.")
    require_owner_or_admin(report.company.owner_id, current_user)

    _check_and_increment_ai_quota(current_user, db)

    def _to_float(val):
        try:
            return float(val) if val is not None else None
        except (TypeError, ValueError):
            return None

    financial_data = {
        "balance_sheet": {
            "total_current_assets": _to_float(report.total_current_assets),
            "total_non_current_assets": _to_float(report.total_non_current_assets),
            "total_assets": _to_float(report.total_assets),
            "total_current_liabilities": _to_float(report.total_current_liabilities),
            "total_non_current_liabilities": _to_float(report.total_non_current_liabilities),
            "total_liabilities": _to_float(report.total_liabilities),
            "total_equity": _to_float(report.total_equity),
            "inventory": _to_float(report.inventory),
            "cash_and_equivalents": _to_float(report.cash_and_equivalents),
        },
        "income_statement": {
            "revenue": _to_float(report.revenue),
            "gross_profit": _to_float(report.gross_profit),
            "ebitda": _to_float(report.ebitda),
            "ebit": _to_float(report.ebit),
            "interest_expense": _to_float(report.interest_expense),
            "net_income": _to_float(report.net_income),
        },
        "cash_flow": {
            "operating_cash_flow": _to_float(report.operating_cash_flow),
            "free_cash_flow": _to_float(report.free_cash_flow),
        },
    }

    try:
        analysis = ai_service.analyze_financial_ratios(financial_data, db, current_user.id)
        report.ai_ratios = analysis
        report.is_ai_generated = True
        notify_user_report_ready(db, current_user.id, report_id, report.company.name)
        db.commit()
        return {"status": "success", "analysis": analysis}
    except Exception as exc:
        user_msg = log_exception(exc, {"report_id": report_id, "user_id": current_user.id})
        raise HTTPException(status_code=500, detail=user_msg)


@router.post("/companies/{company_id}/upload-document", response_model=dict)
async def upload_and_extract(
    company_id: int,
    request: Request,
    file: UploadFile = File(...),
    fiscal_year: int = Form(..., ge=1900, le=2100),
    period: str = Form("annual"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> dict:
    """T3: OCR Pipeline — Belge yükle, AI ile parse et, forma doldur."""
    _get_company_or_403(company_id, current_user, db)
    _validate_fiscal_year(fiscal_year)

    # Content-type header'ı kontrol et (kullanıcı tanımlı — güvenilmez, ama ilk süzgeç)
    if file.content_type not in ALLOWED_MEDIA_TYPES:
        raise HTTPException(status_code=400, detail="Desteklenmeyen dosya türü. PDF, PNG, JPG veya WEBP yükleyin.")

    # Dosya içeriğini oku — önce boyutu doğrula
    content = await file.read()
    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Dosya boyutu 20 MB'ı aşamaz.")
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Boş dosya yüklenemez.")

    # Magic byte doğrulaması — gerçek içerik türünü belirle
    actual_mime = _validate_magic_bytes(content)
    if actual_mime != file.content_type:
        raise HTTPException(status_code=400, detail="Dosya içeriği beyan edilen türle eşleşmiyor.")

    safe_fname = _safe_store_filename(file.filename)

    log_audit(db, current_user.id, LogAction.UPLOAD, "Document", company_id,
              new_values={"file_name": safe_fname, "size": len(content), "mime": actual_mime},
              ip_address=request.client.host if request.client else None)

    _check_and_increment_ai_quota(current_user, db)

    try:
        extracted = ai_service.extract_financial_data_from_document(
            file_path=safe_fname or "upload",
            file_content=content,
            media_type=actual_mime,
            db=db,
            user_id=current_user.id,
        )

        period_val = extracted.get("period", period)
        report_type_val = extracted.get("report_type", "combined")

        try:
            period_enum = PeriodType(period_val)
        except ValueError:
            period_enum = PeriodType.ANNUAL

        try:
            report_type_enum = ReportType(report_type_val)
        except ValueError:
            report_type_enum = ReportType.COMBINED

        bs  = extracted.get("balance_sheet", {})
        inc = extracted.get("income_statement", {})
        cf  = extracted.get("cash_flow", {})

        report = FinancialReport(
            company_id=company_id,
            report_type=report_type_enum,
            period=period_enum,
            fiscal_year=extracted.get("fiscal_year", fiscal_year),
            source_document=safe_fname,
            is_ai_generated=True,
            **{k: v for k, v in bs.items()  if v is not None},
            **{k: v for k, v in inc.items() if v is not None},
            **{k: v for k, v in cf.items()  if v is not None},
        )
        db.add(report)
        db.flush()
        log_audit(db, current_user.id, LogAction.CREATE, "FinancialReport", report.id,
                  new_values={"source": "ai_ocr", "file": safe_fname},
                  ip_address=request.client.host if request.client else None)
        db.commit()
        db.refresh(report)

        return {
            "status": "success",
            "report_id": report.id,
            "confidence_score": extracted.get("confidence_score"),
            "notes": extracted.get("notes"),
            "extracted_data": extracted,
        }

    except HTTPException:
        raise
    except Exception as exc:
        user_msg = log_exception(exc, {"company_id": company_id})
        raise HTTPException(status_code=500, detail=user_msg)


@router.get("/reports/{report_id}/export/pptx")
def export_pptx(
    report_id: int,
    currency: str = Query("TRY", max_length=3),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """T5: Finansal raporu .pptx formatında indir."""
    if currency not in _ALLOWED_CURRENCIES:
        raise HTTPException(status_code=400, detail="Geçersiz para birimi.")

    report = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı.")
    require_owner_or_admin(report.company.owner_id, current_user)

    log_audit(db, current_user.id, LogAction.EXPORT, "FinancialReport", report_id,
              new_values={"format": "pptx"})
    db.commit()

    pptx_bytes = pptx_service.generate_financial_report(
        company=report.company,
        report=report,
        ai_analysis=report.ai_ratios,
        currency=currency,
    )

    raw_name = f"{report.company.name}_{report.fiscal_year}_{report.period.value}.pptx"
    return StreamingResponse(
        io.BytesIO(pptx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": _safe_cd_header(raw_name)},
    )


@router.get("/reports/{report_id}/export/pdf")
def export_pdf(
    report_id: int,
    currency: str = Query("TRY", max_length=3),
    with_ai: bool = Query(True, description="AI anlatı metni ekle (ek AI kotası kullanır)"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Kapsamlı görsel PDF raporu: tablolar, grafikler ve AI yorumları."""
    if currency not in _ALLOWED_CURRENCIES:
        raise HTTPException(status_code=400, detail="Geçersiz para birimi.")

    report = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı.")
    require_owner_or_admin(report.company.owner_id, current_user)

    company_id = report.company_id

    # Genişletilmiş verileri yükle
    banks       = db.query(CompanyBank).filter(CompanyBank.company_id == company_id).all()
    collections = db.query(Collection).filter(Collection.company_id == company_id).all()
    projects    = db.query(CompanyProject).filter(CompanyProject.company_id == company_id).all()
    investments = db.query(Investment).filter(Investment.company_id == company_id).all()

    # AI anlatı metni üret (isteğe bağlı, kota kullanır)
    ai_narrative: Optional[dict] = None
    if with_ai:
        try:
            _check_and_increment_ai_quota(current_user, db)
            def _f(v):
                try:
                    return float(v) if v is not None else None
                except (TypeError, ValueError):
                    return None

            narrative_data = {
                "company": report.company.name,
                "fiscal_year": report.fiscal_year,
                "period": report.period.value,
                "balance_sheet": {
                    "total_assets":               _f(report.total_assets),
                    "total_current_assets":       _f(report.total_current_assets),
                    "total_non_current_assets":   _f(report.total_non_current_assets),
                    "total_current_liabilities":  _f(report.total_current_liabilities),
                    "total_non_current_liabilities": _f(report.total_non_current_liabilities),
                    "total_liabilities":          _f(report.total_liabilities),
                    "total_equity":               _f(report.total_equity),
                    "cash_and_equivalents":       _f(report.cash_and_equivalents),
                    "accounts_receivable":        _f(report.accounts_receivable),
                    "inventory":                  _f(report.inventory),
                },
                "income_statement": {
                    "revenue":          _f(report.revenue),
                    "gross_profit":     _f(report.gross_profit),
                    "ebitda":           _f(report.ebitda),
                    "ebit":             _f(report.ebit),
                    "interest_expense": _f(report.interest_expense),
                    "net_income":       _f(report.net_income),
                },
                "cash_flow": {
                    "operating_cash_flow":  _f(report.operating_cash_flow),
                    "investing_cash_flow":  _f(report.investing_cash_flow),
                    "financing_cash_flow":  _f(report.financing_cash_flow),
                    "free_cash_flow":       _f(report.free_cash_flow),
                },
                "ai_ratios": report.ai_ratios or {},
            }
            ai_narrative = ai_service.generate_report_narrative(
                narrative_data, db, current_user.id
            )
        except HTTPException:
            pass  # kota aşıldıysa anlatısız devam et
        except Exception:
            pass

    log_audit(db, current_user.id, LogAction.EXPORT, "FinancialReport", report_id,
              new_values={"format": "pdf"})
    db.commit()

    try:
        pdf_bytes = report_generator.generate_pdf_report(
            report=report,
            company=report.company,
            banks=banks or None,
            collections=collections or None,
            projects=projects or None,
            investments=investments or None,
            ai_ratios=report.ai_ratios,
            ai_narrative=ai_narrative,
            currency=currency,
        )
    except Exception as exc:
        user_msg = log_exception(exc, {"report_id": report_id})
        raise HTTPException(status_code=500, detail=user_msg)

    raw_name = f"{report.company.name}_{report.fiscal_year}_{report.period.value}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": _safe_cd_header(raw_name)},
    )


@router.delete("/reports/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_report(
    report_id: int,
    request: Request,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> None:
    report = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı.")
    require_owner_or_admin(report.company.owner_id, current_user)

    log_audit(db, current_user.id, LogAction.DELETE, "FinancialReport", report_id,
              old_values={"company_id": report.company_id, "fiscal_year": report.fiscal_year},
              ip_address=request.client.host if request.client else None)
    db.delete(report)
    db.commit()


@router.get("/companies/{company_id}/reports/compare", response_model=dict)
def compare_years(
    company_id: int,
    year_a: int = Query(..., ge=1900, le=2100),
    year_b: int = Query(..., ge=1900, le=2100),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> dict:
    """İki farklı mali yılı yan yana karşılaştırır."""
    _get_company_or_403(company_id, current_user, db)

    def _get_report(year: int) -> Optional[FinancialReport]:
        return (
            db.query(FinancialReport)
            .filter(
                FinancialReport.company_id == company_id,
                FinancialReport.fiscal_year == year,
                FinancialReport.period == PeriodType.ANNUAL,
            )
            .order_by(FinancialReport.created_at.desc())
            .first()
        )

    def _snap(r: Optional[FinancialReport]) -> Optional[dict]:
        if not r:
            return None
        return {
            "report_id": r.id,
            "total_assets":       float(r.total_assets)       if r.total_assets       else None,
            "total_liabilities":  float(r.total_liabilities)  if r.total_liabilities  else None,
            "total_equity":       float(r.total_equity)        if r.total_equity        else None,
            "revenue":            float(r.revenue)             if r.revenue             else None,
            "gross_profit":       float(r.gross_profit)        if r.gross_profit        else None,
            "net_income":         float(r.net_income)          if r.net_income          else None,
            "ebitda":             float(r.ebitda)              if r.ebitda              else None,
            "operating_cash_flow":float(r.operating_cash_flow) if r.operating_cash_flow else None,
            "financial_score":    r.ai_ratios.get("financial_score") if r.ai_ratios else None,
        }

    snap_a, snap_b = _snap(_get_report(year_a)), _snap(_get_report(year_b))

    def _delta(a_val, b_val):
        if a_val is None or b_val is None or b_val == 0:
            return None
        return round((a_val - b_val) / abs(b_val) * 100, 2)

    deltas = {}
    if snap_a and snap_b:
        for key in ("total_assets", "total_liabilities", "total_equity",
                    "revenue", "gross_profit", "net_income", "ebitda", "operating_cash_flow"):
            deltas[key] = _delta(snap_a.get(key), snap_b.get(key))

    return {str(year_a): snap_a, str(year_b): snap_b, "change_pct": deltas}


@router.get("/reports/{report_id}/export/excel")
def export_excel(
    report_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """O6: Finansal raporu Excel (.xlsx) formatında indir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    report = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı.")
    require_owner_or_admin(report.company.owner_id, current_user)

    log_audit(db, current_user.id, LogAction.EXPORT, "FinancialReport", report_id,
              new_values={"format": "excel"})
    db.commit()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rapor {report.fiscal_year}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86AB")
    center = Alignment(horizontal="center")

    headers = ["Kalem", "Değer"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    rows = [
        ("Şirket", report.company.name),
        ("Yıl", report.fiscal_year),
        ("Dönem", report.period.value),
        ("Rapor Türü", report.report_type.value),
        ("", ""),
        ("--- BİLANÇO ---", ""),
        ("Nakit ve Eşdeğerleri", report.cash_and_equivalents),
        ("Alacaklar", report.accounts_receivable),
        ("Stoklar", report.inventory),
        ("Toplam Dönen Varlık", report.total_current_assets),
        ("Toplam Duran Varlık", report.total_non_current_assets),
        ("Toplam Varlık", report.total_assets),
        ("Toplam Kısa Vadeli Yük.", report.total_current_liabilities),
        ("Toplam Uzun Vadeli Yük.", report.total_non_current_liabilities),
        ("Toplam Yükümlülük", report.total_liabilities),
        ("Toplam Öz Kaynak", report.total_equity),
        ("", ""),
        ("--- GELİR TABLOSU ---", ""),
        ("Gelir", report.revenue),
        ("Satış Maliyeti", report.cost_of_goods_sold),
        ("Brüt Kar", report.gross_profit),
        ("EBITDA", report.ebitda),
        ("EBIT", report.ebit),
        ("Faiz Gideri", report.interest_expense),
        ("Net Kar", report.net_income),
        ("", ""),
        ("--- NAKİT AKIŞI ---", ""),
        ("Operasyonel Nakit Akışı", report.operating_cash_flow),
        ("Yatırım Nakit Akışı", report.investing_cash_flow),
        ("Finansman Nakit Akışı", report.financing_cash_flow),
        ("Serbest Nakit Akışı", report.free_cash_flow),
    ]

    for r_idx, (label, value) in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=2, value=float(value) if value is not None and str(value).replace('.','',1).lstrip('-').isdigit() else str(value) if value else None)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    raw_name = f"{report.company.name}_{report.fiscal_year}_{report.period.value}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _safe_cd_header(raw_name)},
    )


@router.get("/reports/{report_id}/export/csv")
def export_csv(
    report_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """O6: Finansal raporu CSV formatında indir."""
    report = (
        db.query(FinancialReport)
        .options(joinedload(FinancialReport.company))
        .filter(FinancialReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı.")
    require_owner_or_admin(report.company.owner_id, current_user)

    log_audit(db, current_user.id, LogAction.EXPORT, "FinancialReport", report_id,
              new_values={"format": "csv"})
    db.commit()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Firma", "Yıl", "Dönem", "Rapor Türü",
        "Toplam Varlık", "Toplam Yükümlülük", "Toplam Özkaynak",
        "Gelir", "Net Kar",
    ])
    writer.writerow([
        report.company.name,
        report.fiscal_year,
        report.period.value,
        report.report_type.value,
        report.total_assets,
        report.total_liabilities,
        report.total_equity,
        report.revenue,
        report.net_income,
    ])
    output.seek(0)

    raw_name = f"{report.company.name}_{report.fiscal_year}_{report.period.value}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": _safe_cd_header(raw_name)},
    )
